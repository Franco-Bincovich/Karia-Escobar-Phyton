"""
Generación y análisis avanzado de archivos Excel (.xlsx).
"""

import os
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from src.utils.app_error import AppError
from src.utils.logger import get_logger

logger = get_logger("excelTools")

TMP_DIR = "/tmp"


async def generar_excel(
    nombreArchivo: str, columnas: list, filas: list, user_id: str, hoja: str = "Datos", **_,
) -> str:
    """
    Genera un archivo Excel (.xlsx) a partir de datos tabulares.

    Returns:
        Ruta del archivo generado en /tmp.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = hoja
        ws.append(columnas)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for col_idx, col_name in enumerate(columnas, 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = max(len(col_name) + 4, 14)

        for fila in filas:
            ws.append(fila)

        ruta = os.path.join(TMP_DIR, f"{user_id}_{nombreArchivo}.xlsx")
        wb.save(ruta)
        logger.info("Excel generado ruta=%s filas=%d", ruta, len(filas))
        return ruta
    except AppError:
        raise
    except Exception as err:
        logger.error("Error al generar Excel: %s", str(err))
        raise AppError(f"Error al generar Excel: {err}", "EXCEL_ERROR", 500)


def _stats_numericas(valores: list[float]) -> dict:
    """Calcula estadísticas para una columna numérica."""
    if not valores:
        return {}
    return {
        "cantidad": len(valores),
        "suma": round(sum(valores), 2),
        "promedio": round(sum(valores) / len(valores), 2),
        "minimo": min(valores),
        "maximo": max(valores),
    }


def _mas_frecuentes(valores: list, n: int = 5) -> list:
    """Devuelve los N valores más frecuentes."""
    return [{"valor": v, "frecuencia": c} for v, c in Counter(valores).most_common(n)]


async def analizar_excel_avanzado(
    nombreArchivo: str, user_id: str, instruccion: str = None, **_,
) -> dict:
    """
    Lee un archivo Excel de /tmp y devuelve análisis estadístico por hoja.

    Returns:
        dict con hojas, resumen_general e instruccion opcional.
    """
    base = nombreArchivo.replace(".xlsx", "").replace(".XLSX", "")
    ruta_con = os.path.join(TMP_DIR, f"{user_id}_{base}.xlsx")
    ruta_sin = os.path.join(TMP_DIR, f"{user_id}_{base}")
    ruta = ruta_con if os.path.exists(ruta_con) else ruta_sin if os.path.exists(ruta_sin) else None

    if not ruta:
        raise AppError(
            f'No se encontró el archivo "{base}.xlsx". Subilo primero.', "ARCHIVO_NO_ENCONTRADO", 404
        )
    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        hojas = []
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                hojas.append({"nombre": nombre_hoja, "filas": 0, "columnas": []})
                continue
            encabezados = [str(c) if c else f"Col{i+1}" for i, c in enumerate(rows[0])]
            cols_data = {i: {"nombre": h, "nums": [], "txts": [], "vacias": 0} for i, h in enumerate(encabezados)}
            for fila in rows[1:]:
                for i, val in enumerate(fila):
                    if i not in cols_data:
                        continue
                    if val is None or val == "":
                        cols_data[i]["vacias"] += 1
                    elif isinstance(val, (int, float)):
                        cols_data[i]["nums"].append(float(val))
                    else:
                        s = str(val).strip()
                        try:
                            cols_data[i]["nums"].append(float(s))
                        except ValueError:
                            cols_data[i]["txts"].append(s)
            columnas = []
            for i in sorted(cols_data):
                c = cols_data[i]
                es_num = len(c["nums"]) > 0 and len(c["nums"]) >= len(c["txts"])
                stats = (
                    {"tipo": "numerico", "vacias": c["vacias"], **_stats_numericas(c["nums"])}
                    if es_num
                    else {"tipo": "texto", "vacias": c["vacias"], "masFrequentes": _mas_frecuentes(c["txts"])}
                )
                columnas.append({"nombre": c["nombre"], "estadisticas": stats})
            hojas.append({"nombre": nombre_hoja, "filas": len(rows) - 1, "columnas": columnas})
        wb.close()
        total_filas = sum(h["filas"] for h in hojas)
        result = {
            "hojas": hojas,
            "resumen_general": {
                "archivo": f"{base}.xlsx", "totalHojas": len(hojas),
                "totalFilas": total_filas,
                "totalColumnas": sum(len(h["columnas"]) for h in hojas),
                "nombresHojas": [h["nombre"] for h in hojas],
            },
        }
        if instruccion:
            result["instruccion"] = instruccion
        logger.info("Excel analizado userId=%s archivo=%s hojas=%d", user_id, base, len(hojas))
        return result
    except AppError:
        raise
    except Exception as err:
        logger.error("Error al analizar Excel: %s", str(err))
        raise AppError(f"Error al analizar el archivo Excel: {err}", "EXCEL_AVANZADO_ERROR", 500)
