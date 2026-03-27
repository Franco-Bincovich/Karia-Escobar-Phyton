"""
Parsea archivos subidos y devuelve su contenido como texto plano.
Soporta: PDF, Excel (.xlsx/.xls), Word (.docx), CSV, TXT.
"""

import csv
import os

from src.utils.app_error import AppError
from src.utils.logger import get_logger

logger = get_logger("documentoService")

LIMITE_CHARS = 80_000

TIPOS_SOPORTADOS = {
    ".pdf": "pdf",
    ".xlsx": "excel",
    ".xls": "excel",
    ".docx": "word",
    ".csv": "csv",
    ".txt": "txt",
}


def _aplicar_limite(texto: str) -> tuple[str, bool]:
    """Trunca el texto si supera el límite y agrega aviso."""
    if len(texto) <= LIMITE_CHARS:
        return texto, False
    truncado = texto[:LIMITE_CHARS]
    truncado += "\n\n[AVISO: El documento fue truncado a 80.000 caracteres por exceder el límite permitido.]"
    return truncado, True


def _parsear_pdf(ruta: str) -> str:
    """Extrae texto de un PDF con pdfplumber."""
    import pdfplumber

    texto = []
    with pdfplumber.open(ruta) as pdf:
        for pagina in pdf.pages:
            t = pagina.extract_text()
            if t:
                texto.append(t)
    return "\n".join(texto)


def _parsear_excel(ruta: str) -> str:
    """Lee todas las hojas de un Excel y las convierte a texto tabular."""
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    lineas = []
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        lineas.append(f"=== Hoja: {nombre_hoja} ===")
        for fila in hoja.iter_rows(values_only=True):
            valores = [str(c) if c is not None else "" for c in fila]
            lineas.append(" | ".join(valores))
    wb.close()
    return "\n".join(lineas)


def _parsear_word(ruta: str) -> str:
    """Extrae texto plano de un Word .docx."""
    from docx import Document

    doc = Document(ruta)
    return "\n".join(p.text for p in doc.paragraphs)


def _parsear_csv(ruta: str) -> str:
    """Lee un CSV y separa columnas con ' | '."""
    lineas = []
    with open(ruta, encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for fila in reader:
            lineas.append(" | ".join(fila))
    return "\n".join(lineas)


def _parsear_txt(ruta: str) -> str:
    """Lee un archivo de texto plano."""
    with open(ruta, encoding="utf-8", errors="replace") as f:
        return f.read()


_PARSERS = {
    "pdf": _parsear_pdf,
    "excel": _parsear_excel,
    "word": _parsear_word,
    "csv": _parsear_csv,
    "txt": _parsear_txt,
}


async def parsear_documento(ruta_archivo: str, nombre_original: str) -> dict:
    """
    Parsea un documento según su extensión y devuelve su texto.

    Args:
        ruta_archivo: Ruta absoluta al archivo temporal.
        nombre_original: Nombre original del archivo (para detectar extensión).

    Returns:
        dict con texto (str), tipo (str), truncado (bool).

    Raises:
        AppError: TIPO_NO_SOPORTADO (415) si la extensión no está soportada.
        AppError: PARSE_ERROR (422) si falla la lectura.
    """
    ext = os.path.splitext(nombre_original)[1].lower()
    tipo = TIPOS_SOPORTADOS.get(ext)

    if not tipo:
        raise AppError(
            f"Tipo de archivo no soportado: {ext}. Formatos válidos: PDF, Excel, Word, CSV, TXT.",
            "TIPO_NO_SOPORTADO",
            415,
        )

    try:
        parser = _PARSERS[tipo]
        texto_raw = parser(ruta_archivo)
        logger.info("Documento parseado nombre=%s tipo=%s chars=%d", nombre_original, tipo, len(texto_raw))
        texto, truncado = _aplicar_limite(texto_raw)
        return {"texto": texto, "tipo": tipo, "truncado": truncado}
    except AppError:
        raise
    except Exception as err:
        logger.error("Error al parsear documento nombre=%s: %s", nombre_original, str(err))
        raise AppError(f"No se pudo procesar el archivo: {err}", "PARSE_ERROR", 422)
    finally:
        if os.path.exists(ruta_archivo):
            try:
                os.unlink(ruta_archivo)
            except OSError as e:
                logger.warning("No se pudo borrar archivo temporal %s: %s", ruta_archivo, str(e))
